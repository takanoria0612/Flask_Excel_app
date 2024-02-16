from flask import Flask, request, redirect, url_for, render_template, send_from_directory, flash, jsonify
import openpyxl
from werkzeug.utils import secure_filename
from collections import defaultdict
from dotenv import load_dotenv
import os
from datetime import datetime, timedelta
from flask_login import LoginManager, login_user, logout_user, login_required, UserMixin
from models import User
from typing import Optional, Dict
import math
from shutil import copyfile
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import requests
import logging

# ログの設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


app = Flask(__name__)
load_dotenv()
app.secret_key = os.getenv('SECRET_KEY')

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # ログインページのビュー関数名

# 環境変数のロード
load_dotenv()

# SMTP設定を環境変数から取得
smtp_server = os.getenv('SMTP_SERVER')
smtp_port = int(os.getenv('SMTP_PORT'))
outlook_email = os.getenv('OUTLOOK_EMAIL')
outlook_password = os.getenv('OUTLOOK_PASSWORD')

# EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH')
EXCEL_FILE_PATH = r"\\DESKTOP-M4FOIJ0\SharedFiles\sales.xlsx"
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


def find_data_by_date(excel_file_path, selected_date_str):
    try:
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active  # ワークブックが正常に読み込まれた後でアクティブなシートを取得
    except FileNotFoundError as e:
        flash("指定されたExcelファイルが見つかりません。", "error")
        logging.error(f"Excelファイルが見つかりません: {e}")
        return None
    except openpyxl.utils.exceptions.InvalidFileException as e:
        flash("無効なExcelファイルです。", "error")
        logging.error(f"無効なExcelファイル: {e}")
        return None
    
    # 選択された日付をdatetimeオブジェクトに変換
    selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    
    # Excelファイルを行ごとにループ
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Excelファイルの日付データがdatetimeオブジェクトとして格納されていると仮定
        row_date = row[0].date() if isinstance(row[0], datetime) else None
        
        # 選択された日付と行の日付が一致するかチェック
        if row_date == selected_date:
            # 一致したら、その行のデータを辞書形式で返す
            data = {
                'exists': True,
                'date': selected_date_str,  # 選択された日付
                'sets': row[1],
                'customers': row[2],
                'bowls': row[3],
                'purchase_total': row[4],
                'total_price': row[5],
                'cash_total': row[6],
                'card_total': row[7],
                'usd_total': row[8],
                'remarks': row[9] if len(row) > 9 else ""  # 備考欄がある場合
            }
            return data
    
    # 一致するデータが見つからなかった場合
    return None

def send_email_with_form_data(form_data):
    """フォームデータをCSV形式でメールで送信する関数"""
    # ヘッダーを追加
    headers = ['日付', '組数', '客数', '丼数', '仕入れ額', '合計値段', '現金合計', 'カード合計', 'USD負担合計' , '備考欄']
    # CSV形式の文字列を作成
    email_body = ",".join(headers) + "\n"

    # フォームデータの値を取得してCSV形式の文字列に変換
    csv_data = ",".join([str(form_data.get(header, '')) for header in [
        'date', 'sets', 'customers', 'bowls', 'purchase_total',
        'total_price', 'cash_total', 'card_total', 'usd_total', 'remarks'
    ]])

    email_body += csv_data

    msg = MIMEText(email_body)
    msg['Subject'] = f"{form_data['date']} 売上集計"
    msg['From'] = outlook_email
    msg['To'] = 'takanori_hiraki@us-design.co.jp'  # 宛先のメールアドレス

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(outlook_email, outlook_password)
            server.sendmail(outlook_email, msg['To'], msg.as_string())
        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False
    
def get_holidays():
    url = os.getenv('HOLIDAYS_API_URL')
    response = requests.get(url)
    holidays = response.json() if response.status_code == 200 else {}
    return holidays

def find_last_business_day(holidays):
    date = datetime.now().date() - timedelta(days=1)  # 前日からスタート
    while date.strftime('%Y-%m-%d') in holidays or date.weekday() >= 5:  # 祝日または土日の場合
        date -= timedelta(days=1)  # さらに1日遡る
    return date

def load_user_from_env() -> Dict[str, User]:
    user_database = {}
    user_count = int(os.getenv('USER_COUNT', 0))  # USER_COUNT 環境変数で管理されるユーザーの数
    for i in range(1, user_count + 1):
        username = os.getenv(f'USERNAME{i}')
        password = os.getenv(f'PASSWORD{i}')
        if username and password:
            user_database[username] = User(str(i), username, password)
    return user_database



@app.route('/set-business-day', methods=['POST'])
def set_business_day():
    try:
        data = request.get_json()
        business_day = data.get('businessDay')
        if not business_day:
            logger.error("Business day not provided in the request.")
            return jsonify({'status': 'error', 'message': 'Business day is required.'}), 400

        # Excelファイルを検索するロジック
        excel_data = find_data_by_date(EXCEL_FILE_PATH, business_day)
        if excel_data:
            flash("前営業日のデータは既にあります", 'info')
            logger.info(f" {business_day} のデータは既に存在します")
            return jsonify({'status': 'success', 'data': excel_data})
        else:
            flash("前営業日のデータはありません。新しいデータを追加してください", 'info')
            logger.info(f"{business_day}のデータは存在しません。追加してください")
            return jsonify({'status': 'not found'})

    except Exception as e:
        logger.error(f"Error while setting business day: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'An error occurred while processing your request.'}), 500




@login_manager.user_loader
def load_user(user_id: str) -> Optional[User]:
    # ユーザーデータベースを更新
    user_database = load_user_from_env()
    
    # user_idに対応するユーザーを探す
    user = next((user for user in user_database.values() if user.id == user_id), None)
    return user



# ログインページ

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user_database = load_user_from_env() 
        user = user_database.get(username, None)
        if user and user.password == password:
            login_user(user)
            return redirect(url_for('index'))
        else:
            return 'ユーザー名、またはパスワードが正しくありません。'
    return render_template('login.html')


# ログアウトページ
@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/')
@login_required
def index():
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    # データの初期化
    data = []
    total_price = 0.0
    total_purchase = 0.0

    # Excelファイルの存在をチェックし、結果を変数に格納
    file_exists = os.path.exists(EXCEL_FILE_PATH)

    if file_exists:
        try:
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active

            # Excelファイルからデータを読み込む
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if isinstance(row[0], datetime):
                    row_date = row[0].date()
                else:
                    row_date = datetime.strptime(str(row[0]), "%Y-%m-%d").date()

                if row_date.year == current_year and row_date.month == current_month:
                    # 合計値段と客数を取得
                    total = float(row[4]) if row[4] else 0
                    customers = int(row[2]) if row[2] else 0
                    purchase = float(row[8]) if row[8] else 0
                    # 客単価を計算
                    avg_spend = total / customers if customers > 0 else 0
                    # データリストに行と客単価を追加
                    data.append(row + (avg_spend,))
                    total_price += total
                    total_purchase += purchase
            total_price = int(total_price)
            total_purchase = int(total_purchase)

        except Exception as e:
            logging.error(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
            flash('Excelファイルの読み込み中にエラーが発生しました。', 'error')
            file_exists = False  # ここでfile_existsを更新してはいけません
    else:
        error_message = "Excelファイルが見つかりません"
        logging.error(error_message)
        flash(error_message, "error_index")

    # file_exists の状態に関わらず、テンプレートに必要な変数を渡す
    return render_template('index.html', file_exists=file_exists, data=data, total_price=total_price, total_purchase=total_purchase)

# @app.route('/')
# @login_required
# def index():
#     now = datetime.now()
#     current_year = now.year
#     current_month = now.month
#     # Excelファイルの存在をチェックし、結果を変数に格納
#     file_exists = os.path.exists(EXCEL_FILE_PATH)
#     # Excelファイルの存在をチェック
#     if not os.path.exists(EXCEL_FILE_PATH):
#         error_message = f"Excelファイルが見つかりません"
#         logging.error(error_message)  # ログにエラーメッセージを記録
#         flash(error_message, "error_index")  # ユーザーにエラーメッセージを表示
#         return render_template('index.html', file_exists=file_exists) 
#     else:
#         try:
#             # Excelファイルを開く
#             workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
#             sheet = workbook.active
#         except Exception as e:
#             logging.error(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
#             flash('Excelファイルの読み込み中にエラーが発生しました。', 'error')



#     data = []
#     total_price = 0.0
#     total_purchase = 0.0
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if isinstance(row[0], datetime):
#             row_date = row[0].date()
#         else:
#             row_date = datetime.strptime(str(row[0]), "%Y-%m-%d").date()

#         if row_date.year == current_year and row_date.month == current_month:
#             # 合計値段と客数を取得
#             total = float(row[4]) if row[4] else 0
#             customers = int(row[2]) if row[2] else 0
#             purchase = float(row[8]) if row[8] else 0  # Get purchase amount
#             # 客単価を計算 (0で除算しないようにする)
#             avg_spend = total / customers if customers > 0 else 0
#             # データリストに行と客単価を追加
#             data.append(row + (avg_spend,))
#             total_price += total
#             # 小数点以下を表示しない形式でフォーマットする
#             total_purchase += purchase  # Add to total purchase
#     total_price = int(total_price)
#     total_purchase = int(total_purchase)  # Convert total purchase to an integer
#     return render_template('index.html', file_exists=file_exists, data=data, total_price=total_price, total_purchase=total_purchase)


@app.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    # 初期値として昨日の日付を設定
    default_date = (datetime.now() - timedelta(days=1)).date()

    # 初期フォームデータの設定
    form_data = {
        # 'date': default_date.strftime('%Y-%m-%d'),  # 昨日の日付をデフォルト値として設定
        'sets': '',
        'customers': '',
        'bowls': '',
        'purchase_total': '',
        'total_price': '',
        'cash_total': '',
        'card_total': '',
        'usd_total': '',
        'remarks': '',
        # 'per_customer_price': ''
    }
    if request.method == 'GET':
        business_day = request.args.get('businessDay')  # フロントエンドから送信されたbusinessDayを取得
        # 新たに客単価を計算
        # 'customers' フィールドの検証と変換
        customers_str = form_data.get('customers', '').strip()
        customers = int(customers_str) if customers_str.isdigit() else 1  # 数字のみの場合に変換、それ以外はデフォルト値1
        total_price_str = form_data.get('total_price', '').strip()
        total_price = float(total_price_str) if total_price_str.replace('.', '', 1).isdigit() else 0

        per_customer_price = total_price / customers if customers > 0 else 0
    elif request.method == 'POST':
        form_data = request.form.to_dict(flat=True)  # Use flat=True to get a regular dict
        # form_dataから日付文字列を取得し、datetimeオブジェクトに変換後、dateオブジェクトに変換
        date_str = form_data['date']
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()  # 日付のみを含むdateオブジェクト



        # 新たに客単価を計算
        customers = int(form_data.get('customers', 1))  # 0除算を避けるためデフォルトを1に
        total_price = float(form_data.get('total_price', 0))
        per_customer_price = total_price / customers if customers > 0 else 0

        # Excelファイル内のデータをチェックし、更新する行を探す
        try:
            update_row = None
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_date = row[0].value
                if isinstance(row_date, datetime):
                    row_date = row_date.date()
                if row_date == date_obj:
                    update_row = row
                    break

            if update_row:  # 既存のデータを更新
                # 各セルに新しい値を設定
                update_row[1].value = int(form_data.get('sets', 0))
                update_row[2].value = int(form_data.get('customers', 0))
                update_row[3].value = int(form_data.get('bowls', 0))
                update_row[4].value = float(form_data.get('purchase_total', 0))
                update_row[5].value = float(form_data.get('total_price', 0))
                update_row[6].value = float(form_data.get('cash_total', 0))
                update_row[7].value = float(form_data.get('card_total', 0))
                update_row[8].value = float(form_data.get('usd_total', 0))
                update_row[9].value = form_data.get('remarks', '')
                # update_row[10].value = float(per_customer_price) 
                # update_row[10].value = float(form_data.get('per_customer_price', 3)) # 客単価を新たな列に設定
                flash('データを更新しました。', 'success')
            else:  # 新しいデータを追加
                sheet.append([
                    date_obj,
                    int(form_data.get('sets', 0)),
                    int(form_data.get('customers', 0)),
                    int(form_data.get('bowls', 0)),
                    float(form_data.get('purchase_total', 0)),
                    float(form_data.get('total_price', 0)),
                    float(form_data.get('cash_total', 0)),
                    float(form_data.get('card_total', 0)),
                    float(form_data.get('usd_total', 0)),
                    form_data.get('remarks', ''),
                    # float(per_customer_price)  # 客単価を追加
                    # float(form_data.get('per_customer_price', 2))
                ])
                flash('新しいデータを追加しました。', 'success')

            workbook.save(EXCEL_FILE_PATH)



            email_sent = send_email_with_form_data(form_data)
            if email_sent:
                flash('メールを送信しました。','success')
                return redirect(url_for('add'))  # メール送信後に適切なページにリダイレクト
            else:
                flash('メールの送信に失敗しました。', 'error')

        except ValueError as e:
            # ここでエラーメッセージとともにフォームページにリダイレクト
            flash(str(e), 'error')
            # ここでエラーメッセージとともにフォームページに値を渡してレンダリング
            return render_template('add.html', form_data=form_data)
        
        

    # add.htmlを表示
    return render_template('add.html', form_data=form_data)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
    print('file doesnt exist')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 許可されたファイルの拡張子をチェックする関数
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/fetch-data-for-date')
@login_required
def fetch_data_for_date():
    selected_date = request.args.get('date')
    data = find_data_by_date(EXCEL_FILE_PATH, selected_date)
    if data:
        return jsonify(data)
    else:
        return jsonify({'exists': False})

# @app.route('/upload', methods=['GET', 'POST'])
# @login_required
# def upload_file():
#     if request.method == 'POST':
#         if 'file' not in request.files:
#             flash('ファイルが選択されていません。')
#             return redirect(request.url)
#         file = request.files['file']
#         if file.filename == '':
#             flash('ファイル名がありません。')
#             return redirect(request.url)
#         if file and allowed_file(file.filename):
#             if file.filename != 'sales.xlsx':
#                 flash('アップロードされたファイル名が正しくありません。sales.xlsxという名前のファイルをアップロードしてください。')
#                 return redirect(request.url)

#             # アップロードされたファイルを一時的に保存
#             temp_filename = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
#             file.save(temp_filename)
            
#             workbook = openpyxl.load_workbook(temp_filename)
#             sheet = workbook.active
            
#             # データ加工処理
#             for row in sheet.iter_rows(min_row=2, values_only=False):
#                 if row[9].value is None:
#                     total_sales = float(row[4].value) or 0
#                     customer_count = int(row[2].value) or 0
#                     row[9].value = total_sales / customer_count if customer_count else 0
            
#             # 加工後の内容を指定の共有ディレクトリに保存
#             final_path = r"\\DESKTOP-M4FOIJ0\SharedFiles\sales.xlsx"
#             workbook.save(final_path)

#             flash('ファイルが正常にアップロードされ、処理されました。')
#             return redirect(url_for('upload_file'))
#         else:
#             flash('このファイルタイプは許可されていません。')
#             return redirect(request.url)
#     return render_template('upload.html')


# @app.route('/download')
# @login_required
# def download_file():
#     directory = os.path.dirname(EXCEL_FILE_PATH)
#     filename = os.path.basename(EXCEL_FILE_PATH)
#     return send_from_directory(directory, filename, as_attachment=True)


@app.route('/filter', methods=['GET'])
@login_required
def filter_data():
    selected_month = request.args.get('selectedMonth', '')  # デフォルト値を空文字列に設定
    file_exists = os.path.exists(EXCEL_FILE_PATH)  # ファイルの存在を確認

    if not file_exists:
        flash("Excelファイルが見つかりません。", "error")
        return redirect(url_for('index'))

    year, month = map(int, selected_month.split('-'))

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    except Exception as e:
        flash("Excelファイルを開く際にエラーが発生しました。", "error")
        return redirect(url_for('index'))

    filtered_data = []
    total_purchase = 0.0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], datetime):
            row_date = row[0]
        else:
            try:
                row_date = datetime.strptime(row[0], "%Y-%m-%d")
            except ValueError:
                continue

        if row_date.year == year and row_date.month == month:
            total_sales = float(row[4]) if row[4] else 0
            customer_count = int(row[2]) if row[2] else 0
            purchase_amount = float(row[8]) if row[8] else 0
            total_purchase += purchase_amount
            average_spend_per_customer = total_sales / customer_count if customer_count > 0 else 0
            row_data_with_avg_spend = list(row) + [average_spend_per_customer]
            filtered_data.append(row_data_with_avg_spend)

    total_purchase = int(total_purchase)
    total_price = sum(row[4] for row in filtered_data if row[4])

    return render_template('index.html', file_exists=True, data=filtered_data, total_price=total_price, total_purchase=total_purchase, selected_month=selected_month)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5500, debug=True)
